import os
import math
import locale
from datetime import datetime, time
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
import getpass

def get_windows_user():
    return getpass.getuser()

user = get_windows_user()
# Define local para português (Brasil)
locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')

def gerar_planilha_para_cliente(cliente_nome, dados_cliente_lista):
    wb = load_workbook("C:/Users/raul.araujo/Downloads/MODELO MEDIÇÃO 2025.xlsx")
    ws = wb.active

    linha_atual = 10
    total_valor_desconto = 0

    chaves_ordenadas = [
        'cliente', 'Placa 1', 'Placa 2', 'Chassi', 'AIT', 'Data Da Infração', 
        'Hora Da Infração', 'Cód. Da Infração', 'Descrição Da Infração', 
        'Local Da Infração', 'Taxa ADM', 'Valor com Desconto', 'Valor da taxa ADM','Valor total Reembolsável '
    ]

    for dados_cliente in dados_cliente_lista:
        col = 1
        for chave in chaves_ordenadas:
            valor = dados_cliente.get(chave, "")
            if isinstance(valor, float) and math.isnan(valor):
                valor = ""

            # Corrigir soma
            if chave == 'Valor total Reembolsável ':
                try:
                    if isinstance(valor, str):
                        valor = valor.replace("R$", "").replace(".", "").replace(",", ".").strip()
                    valor_float = float(valor)
                    total_valor_desconto += valor_float
                except:
                    pass

            # Datas e horários
            if chave == 'Hora Da Infração':
                if isinstance(valor, time):
                    valor = valor.strftime('%H:%M')
                elif isinstance(valor, float):
                    try:
                        valor = from_excel(valor).strftime('%H:%M')
                    except:
                        valor = ""
            elif chave == 'Data Da Infração':
                if isinstance(valor, float):
                    try:
                        valor = from_excel(valor).strftime('%d/%m/%Y')
                    except:
                        valor = ""
                elif isinstance(valor, datetime):
                    valor = valor.strftime('%d/%m/%Y')
            elif valor is None:
                valor = ""

            ws.cell(row=linha_atual, column=col).value = valor
            ws.cell(row=linha_atual, column=col).alignment = ws.cell(row=linha_atual, column=col).alignment.copy(horizontal='center')
            col += 1
        
        linha_atual += 1
    ws['D7'] = f"Valor Faturado: R$ {total_valor_desconto:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".")
    ws['B4'] = cliente_nome
    cnpj = dados_cliente_lista[0].get('CNPJ/CPF', 'N/A')
    ws['B5'] = f"CNPJ: {cnpj}"


    mes_ano = datetime.now().strftime('%B/%Y').capitalize()
    ws['B6'] = mes_ano

    
    qtd_multas = len(dados_cliente_lista)
    ws['B7'] = f"QTD. MULTAS: {qtd_multas}"
    # Salvar
    caminho_saida = f"C:/Users/raul.araujo/Downloads/MEDICAO_{cliente_nome}.xlsx"
    wb.save(caminho_saida)
    return caminho_saida


from openpyxl import load_workbook
import pandas as pd

def extrair_links_reais(path_arquivo, coluna_link='Link'):
    # Carrega o DataFrame e o workbook
    df = pd.read_excel(path_arquivo)
    wb = load_workbook(path_arquivo, data_only=True)
    ws = wb.active
    # Encontra o índice da coluna do link
    col_index = None
    for idx, cell in enumerate(ws[1]):
        if str(cell.value).strip().lower() == coluna_link.lower():
            col_index = idx
            break
    if col_index is None:
        raise ValueError(f"Coluna '{coluna_link}' não encontrada.")

    # Substitui o conteúdo da coluna pelos hyperlinks reais
    links = []
    for row in ws.iter_rows(min_row=2):
        cell = row[col_index]
        if cell.hyperlink:
            links.append(cell.hyperlink.target)
        else:
            links.append(None)

    df[coluna_link] = links
    return df
